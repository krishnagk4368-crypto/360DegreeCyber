# backend/app/services/excel_service.py
from io import BytesIO
from typing import Iterable
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

def findings_to_xlsx(rows: Iterable) -> BytesIO:
    """
    rows: iterable of Finding ORM objects (with attributes:
          id, project_id, title, severity, status, description, poc_path)
    returns a BytesIO ready to stream.
    """
    wb = Workbook()
    ws = wb.active
    ws.title = "Findings"

    headers = ["id", "project_id", "title", "severity", "status", "description", "poc_path"]
    ws.append(headers)

    for f in rows:
        ws.append([
            f.id, f.project_id, f.title, f.severity, f.status,
            f.description or "", f.poc_path or ""
        ])

    # Auto-fit-ish: set width based on content length (simple heuristic)
    for col_idx, col in enumerate(ws.columns, start=1):
        max_len = max((len(str(c.value)) if c.value else 0) for c in col)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(max(12, max_len + 2), 60)

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf
